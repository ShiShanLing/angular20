"""
monthly_dca_report.py

用途：
- 每月跑一次 ETF / 指数定投回测
- 自动下载价格数据（默认用 yfinance）
- 模拟“每月定投一次，持有到现在”
- 输出 Excel：汇总表 + 每个资产的明细

安装：
pip install pandas numpy yfinance openpyxl

运行：
python monthly_dca_report.py

输出：
- 默认写入本仓库 reports/monthly_dca_report.xlsx（与脚本同级目录下的 reports 文件夹）
- 各工作表自动生成图表：汇总为各资产 5年/10年 XIRR 对比；基准参考为外部五年/十年年化；明细表为「起始日×年化XIRR」柱状图
- 汇总表与「指数基准参考」表含估值列（Yahoo 当前PE、推算的十年平均PE、倍数百分比）；基准表需在 BENCHMARK_INDEX_PE_TICKER 配置拉取代码
"""

from __future__ import annotations

import math
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import yfinance as yf

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# 你主要改这里
# =========================

ASSETS: Dict[str, Dict[str, str]] = {
    # 美股 ETF（benchmark_key 对应 BENCHMARK_TABLE 中的「指数名称」，无则不在汇总中填参考列）
    "SPY": {"ticker": "SPY", "category": "美股", "benchmark_key": "SPY (S&P 500)"},
    "QQQ": {"ticker": "QQQ", "category": "美股", "benchmark_key": "QQQ (Nasdaq 100)"},
    "VT": {"ticker": "VT", "category": "美股", "benchmark_key": "VT (Total World)"},
    "SCHD": {"ticker": "SCHD", "category": "美股", "benchmark_key": "SCHD"},
    "GLD": {"ticker": "GLD", "category": "黄金", "benchmark_key": "GLD (Gold)"},

    # A股 ETF 示例：需要确认代码是否与你使用的平台一致
    # yfinance 常用格式：沪市 .SS，深市 .SZ
    "沪深300ETF": {"ticker": "510300.SS", "category": "宽基", "benchmark_key": "沪深300"},
    "中证500ETF": {"ticker": "510500.SS", "category": "宽基", "benchmark_key": "中证500"},
    "创业板ETF": {"ticker": "159915.SZ", "category": "成长", "benchmark_key": "创业板"},
}


# ---------------------------------------------------------------------------
# 市场指数参考统计（来自你提供的对照表截图；列为「外部参考表」上的 5 年 / 10 年年化，与 yfinance 回测独立）
# 数值为小数（如 0.09 = 9%）；无数据用 None，导出为 Excel 空单元格。
# ---------------------------------------------------------------------------
BENCHMARK_TABLE: List[Dict[str, Any]] = [
    # 红利/价值
    {"分类": "红利/价值", "指数名称": "红利100", "外部参考表_五年年化": 0.120, "外部参考表_十年年化": 0.097},
    {"分类": "红利/价值", "指数名称": "富国红利", "外部参考表_五年年化": 0.094, "外部参考表_十年年化": 0.081},
    # 宽基旗舰
    {"分类": "宽基旗舰", "指数名称": "沪深300", "外部参考表_五年年化": 0.065, "外部参考表_十年年化": 0.063},
    {"分类": "宽基旗舰", "指数名称": "A500", "外部参考表_五年年化": 0.075, "外部参考表_十年年化": 0.075},
    {"分类": "宽基旗舰", "指数名称": "A50", "外部参考表_五年年化": 0.055, "外部参考表_十年年化": 0.073},
    {"分类": "宽基旗舰", "指数名称": "上证50", "外部参考表_五年年化": 0.058, "外部参考表_十年年化": 0.046},
    {"分类": "宽基旗舰", "指数名称": "中证500", "外部参考表_五年年化": 0.057, "外部参考表_十年年化": 0.078},
    {"分类": "宽基旗舰", "指数名称": "中证1000", "外部参考表_五年年化": 0.041, "外部参考表_十年年化": 0.059},
    # 成长
    {"分类": "成长", "指数名称": "创业板", "外部参考表_五年年化": 0.071, "外部参考表_十年年化": 0.128},
    # 美股（美金计价）
    {"分类": "美股", "指数名称": "SCHD", "外部参考表_五年年化": 0.117, "外部参考表_十年年化": 0.124},
    {"分类": "美股", "指数名称": "SPY (S&P 500)", "外部参考表_五年年化": 0.090, "外部参考表_十年年化": 0.140},
    {"分类": "美股", "指数名称": "VT (Total World)", "外部参考表_五年年化": 0.097, "外部参考表_十年年化": 0.134},
    {"分类": "美股", "指数名称": "GLD (Gold)", "外部参考表_五年年化": 0.048, "外部参考表_十年年化": 0.181},
    {"分类": "美股", "指数名称": "XLE (Energy)", "外部参考表_五年年化": 0.072, "外部参考表_十年年化": 0.105},
    {"分类": "美股", "指数名称": "QQQ (Nasdaq 100)", "外部参考表_五年年化": 0.172, "外部参考表_十年年化": 0.190},
    # 行业主题
    {"分类": "行业主题", "指数名称": "全指消费", "外部参考表_五年年化": 0.133, "外部参考表_十年年化": 0.010},
    {"分类": "行业主题", "指数名称": "全指医药", "外部参考表_五年年化": 0.106, "外部参考表_十年年化": 0.002},
    {"分类": "行业主题", "指数名称": "养老产业", "外部参考表_五年年化": -0.021, "外部参考表_十年年化": -0.020},
    {"分类": "行业主题", "指数名称": "全指信息", "外部参考表_五年年化": 0.095, "外部参考表_十年年化": 0.101},
    {"分类": "行业主题", "指数名称": "环保", "外部参考表_五年年化": 0.025, "外部参考表_十年年化": 0.068},
    {"分类": "行业主题", "指数名称": "传媒", "外部参考表_五年年化": -0.018, "外部参考表_十年年化": 0.101},
    {"分类": "行业主题", "指数名称": "全指金融", "外部参考表_五年年化": 0.080, "外部参考表_十年年化": 0.069},
    {"分类": "行业主题", "指数名称": "中概互联", "外部参考表_五年年化": None, "外部参考表_十年年化": -0.007},
    # 港股
    {"分类": "港股", "指数名称": "恒生ETF", "外部参考表_五年年化": -0.014, "外部参考表_十年年化": 0.054},
    {"分类": "港股", "指数名称": "港股通", "外部参考表_五年年化": None, "外部参考表_十年年化": 0.182},
]

# 「指数基准参考」表中指数名称 -> 用于拉 PE 的 Yahoo 代码（多为同名 ETF/指数基金，非官方指数 PE；可自行改）
# 未列出的行不拉 PE，估值列为空
BENCHMARK_INDEX_PE_TICKER: Dict[str, str] = {
    "沪深300": "510300.SS",
    "中证500": "510500.SS",
    "中证1000": "512100.SS",
    "创业板": "159915.SZ",
    "上证50": "510050.SS",
    # A 系列宽基常见 ETF（上交所一律用 .SS；Yahoo 不认 .SH）
    # A500：563220 若雅虎无数据可改 512050.SS（同中证A500指数）
    "A500": "512050.SS",
    "A50": "560050.SS",
    "红利100": "515100.SS",  # 示例：红利指数基金
    "SPY (S&P 500)": "SPY",
    "QQQ (Nasdaq 100)": "QQQ",
    "SCHD": "SCHD",
    "VT (Total World)": "VT",
    "GLD (Gold)": "GLD",
    "XLE (Energy)": "XLE",
    "恒生ETF": "159920.SZ",
    "中概互联": "513050.SS",
    # 港股通与「中概互联」不同标的；示例：港股通50 ETF（可改为你用的港股通宽基代码）
    "港股通": "513550.SS",
}

START_DATE = "2010-01-01"
MONTHLY_INVEST = 1000

# yfinance 经 Yahoo 拉数，易出现瞬时网络/限流/SSL 失败；首轮失败后按资产分轮重试
DOWNLOAD_MAX_RETRIES = 3

# 明细表「定投起始日 × 年化XIRR」柱状图：点数过多时只画时间上最后 N 行
DCA_BAR_CHART_MAX_POINTS = 180

# 与 monthly_dca_report.py 同级的 angular20/reports/ 下，避免随终端 cwd 写到别处
_PROJECT_DIR = Path(__file__).resolve().parent
OUTPUT_FILE = _PROJECT_DIR / "reports" / "monthly_dca_report.xlsx"

# 各资产「明细」工作表：导出为中文表头（含义与 run_dca_to_now 列一致）
RUNS_DETAIL_COLUMNS_ZH: Dict[str, str] = {
    "start_date": "定投起始日",
    "end_date": "回测截止日(最新月末)",
    "years": "持有年数(至截止日)",
    "months": "定投月数",
    "total_invest": "累计投入",
    "final_value": "期末市值",
    "total_return": "累计回报率(该路径)",
    "xirr": "年化XIRR(该路径)",
    "max_loss": "最大回撤(该路径)",
}


def benchmark_reference_frame() -> pd.DataFrame:
    """指数基准表 + 对映射到的 Yahoo标 的拉取 PE（与汇总列名一致）。"""
    rows_out: List[Dict[str, Any]] = []
    for r in BENCHMARK_TABLE:
        row = dict(r)
        name = str(row["指数名称"])
        ysym = BENCHMARK_INDEX_PE_TICKER.get(name)
        row["估值_PE拉取用代码"] = ysym or ""
        if not ysym:
            row["估值_当前PE"] = None
            row["估值_十年平均PE"] = None
            row["估值_当前相对十年均值_百分比"] = None
            row["估值_数据说明"] = "未配置PE代码(见脚本 BENCHMARK_INDEX_PE_TICKER)"
        else:
            vm = valuation_metrics_for_ticker(ysym)
            row["估值_当前PE"] = vm["估值_当前PE"]
            row["估值_十年平均PE"] = vm["估值_十年平均PE"]
            row["估值_当前相对十年均值_百分比"] = vm["估值_当前相对十年均值_百分比"]
            vn = vm.get("估值_数据说明") or ""
            row["估值_数据说明"] = (f"PE标的:{ysym}" + ("；" + vn if vn else "")).strip("；")
        rows_out.append(row)
    return pd.DataFrame(rows_out)


def benchmark_yield_lookup() -> Dict[str, Tuple[Optional[float], Optional[float]]]:
    """指数名称 -> (外部参考五年年化, 外部参考十年年化)，可能为 None。"""
    out: Dict[str, Tuple[Optional[float], Optional[float]]] = {}
    for r in BENCHMARK_TABLE:
        name = str(r["指数名称"])
        out[name] = (r["外部参考表_五年年化"], r["外部参考表_十年年化"])
    return out


def attach_benchmark_columns(
    row: Dict[str, Any],
    info: Dict[str, str],
    lookup: Dict[str, Tuple[Optional[float], Optional[float]]],
) -> None:
    """附加「外部参考表」年化，及与「持有期约5/10年」回测 XIRR 中位数的差额（若有）。"""
    key = info.get("benchmark_key")
    if not key:
        return
    pair = lookup.get(key)
    if not pair:
        return
    y5, y10 = pair
    if y5 is not None:
        row["外部参考表_五年年化"] = y5
    if y10 is not None:
        row["外部参考表_十年年化"] = y10

    sim5 = row.get("持有期约5年_年化XIRR_中位数")
    sim10 = row.get("持有期约10年_年化XIRR_中位数")
    if isinstance(sim5, (int, float)) and y5 is not None:
        row["持有期约5年_相对外部参考_中位差额(回测减参考)"] = float(sim5) - y5
    if isinstance(sim10, (int, float)) and y10 is not None:
        row["持有期约10年_相对外部参考_中位差额(回测减参考)"] = float(sim10) - y10


# =========================
# 工具函数
# =========================

def xirr(cashflows: List[tuple[pd.Timestamp, float]], guess: float = 0.08) -> Optional[float]:
    """
    简易 XIRR。
    cashflows: [(date, amount)]
    投入为负数，期末市值为正数。
    """
    if not cashflows:
        return None

    dates = [pd.Timestamp(d) for d, _ in cashflows]
    amounts = [float(v) for _, v in cashflows]
    d0 = dates[0]

    def npv(rate: float) -> float:
        total = 0.0
        for d, a in zip(dates, amounts):
            years = (d - d0).days / 365.25
            total += a / ((1 + rate) ** years)
        return total

    rate = guess
    for _ in range(100):
        f = npv(rate)
        eps = 1e-6
        try:
            derivative = (npv(rate + eps) - f) / eps
            if abs(derivative) < 1e-10:
                break
            new_rate = rate - f / derivative
            if not math.isfinite(new_rate) or new_rate <= -0.999:
                break
            if abs(new_rate - rate) < 1e-8:
                return new_rate
            rate = new_rate
        except Exception:
            break

    # 二分兜底
    low, high = -0.99, 10.0
    try:
        for _ in range(200):
            mid = (low + high) / 2
            if npv(low) * npv(mid) <= 0:
                high = mid
            else:
                low = mid
        return (low + high) / 2
    except Exception:
        return None


def download_monthly_price(ticker: str, start: str) -> pd.DataFrame:
    raw = yf.download(ticker, start=start, auto_adjust=True)

    if raw.empty:
        raise ValueError(f"下载失败或无数据：{ticker}")

    if isinstance(raw.columns, pd.MultiIndex):
        price = raw["Close"].iloc[:, 0]
    else:
        price = raw["Close"]

    df = price.to_frame("price").dropna()
    df.index = pd.to_datetime(df.index)
    df = df.resample("ME").last().dropna()
    df = df.reset_index().rename(columns={"Date": "date"})
    return df


def _pick_quarterly_eps_series(q: pd.DataFrame) -> Optional[pd.Series]:
    for label in q.index:
        sl = str(label).lower()
        if "diluted" in sl and "eps" in sl:
            return q.loc[label]
    for label in q.index:
        sl = str(label).lower()
        if "basic" in sl and "eps" in sl:
            return q.loc[label]
    for key in ("Diluted EPS", "Basic EPS"):
        if key in q.index:
            return q.loc[key]
    return None


def _rolling_avg_pe_from_quarterly_eps(tk: Any, hist: pd.DataFrame) -> Optional[float]:
    """
    用最近四个季度 Diluted EPS 之和近似 TTM EPS，对每个财报季末取价算 PE，
    再对最近至多 40 个季度点取平均，作为「十年平均 PE」近似。
    """
    q = getattr(tk, "quarterly_income_stmt", None)
    if q is None or q.empty:
        return None
    eps_row = _pick_quarterly_eps_series(q)
    if eps_row is None:
        return None
    try:
        cols = sorted(q.columns, key=lambda c: pd.Timestamp(c))
    except Exception:
        return None
    pes: List[float] = []
    for i in range(3, len(cols)):
        try:
            ttm = 0.0
            ok = True
            for j in range(i - 3, i + 1):
                v = eps_row[cols[j]]
                if v is None or pd.isna(v):
                    ok = False
                    break
                ttm += float(v)
            if not ok or ttm <= 0 or not math.isfinite(ttm):
                continue
            dt = pd.Timestamp(cols[i])
            if pd.isna(dt):
                continue
            if dt.tzinfo is not None:
                dt = dt.tz_localize(None)
            window = hist.loc[: dt]
            if window.empty:
                continue
            px = float(window["Close"].iloc[-1])
            if px <= 0:
                continue
            pes.append(px / ttm)
        except Exception:
            continue
    if len(pes) < 4:
        return None
    tail = pes[-40:] if len(pes) > 40 else pes
    return float(np.mean(tail))


def _pick_annual_eps_series(inc: pd.DataFrame) -> Optional[pd.Series]:
    """从 yfinance 年报 income_stmt 中取 EPS 行（优先 Diluted）。"""
    for label in inc.index:
        sl = str(label).lower()
        if "diluted" in sl and "eps" in sl:
            return inc.loc[label]
    for label in inc.index:
        sl = str(label).lower()
        if "basic" in sl and "eps" in sl:
            return inc.loc[label]
    for key in ("Diluted EPS", "Basic EPS"):
        if key in inc.index:
            return inc.loc[key]
    return None


def _valuation_fill_percentage_vs_avg(out: Dict[str, Any]) -> None:
    """当前 PE / 十年平均 PE × 100；>100 表示高于十年简单均值。"""
    c = out.get("估值_当前PE")
    a = out.get("估值_十年平均PE")
    if c is not None and a is not None and float(a) > 0:
        out["估值_当前相对十年均值_百分比"] = round(float(c) / float(a) * 100.0, 2)


def valuation_metrics_for_ticker(ticker: str) -> Dict[str, Any]:
    """
    Yahoo / yfinance 估值快照：
    - 估值_当前PE：优先 trailingPE，否则 forwardPE（来自 Yahoo 快照，部分标的为空）
    - 估值_十年平均PE：优先年报 EPS×历史价；不足时用季报近四季 EPS 之和近似 TTM 再滚动（仍为近似）
    - 估值_当前相对十年均值_百分比：当前 / 十年平均 × 100
    ETF、指数、黄金等若 Yahoo 不给 PE 或缺少季报，则相应为空——不是您理解错了，是数据源有限。
    """
    out: Dict[str, Any] = {
        "估值_当前PE": None,
        "估值_十年平均PE": None,
        "估值_当前相对十年均值_百分比": None,
        "估值_数据说明": "",
    }
    notes: List[str] = []
    try:
        tk = yf.Ticker(ticker)
        info = getattr(tk, "info", None) or {}

        cur: Optional[float] = None
        for key in ("trailingPE", "forwardPE"):
            raw = info.get(key)
            if raw is None:
                continue
            try:
                v = float(raw)
                if math.isfinite(v) and v > 0:
                    cur = v
                    if key == "forwardPE":
                        notes.append("当前PE为forwardPE")
                    break
            except (TypeError, ValueError):
                continue
        if cur is not None:
            out["估值_当前PE"] = round(cur, 3)

        hist = tk.history(period="12y", interval="1d", auto_adjust=True)
        if hist.empty:
            if notes:
                out["估值_数据说明"] = "；".join(notes)
            _valuation_fill_percentage_vs_avg(out)
            return out

        if getattr(hist.index, "tz", None) is not None:
            hist = hist.copy()
            hist.index = hist.index.tz_localize(None)

        inc = getattr(tk, "income_stmt", None)
        annual_pes: List[float] = []
        if inc is not None and not inc.empty:
            eps_series = _pick_annual_eps_series(inc)
            if eps_series is None:
                notes.append("年报无可用EPS行")
            else:
                for col in inc.columns:
                    try:
                        eps_val = eps_series[col]
                        if eps_val is None or pd.isna(eps_val):
                            continue
                        eps_f = float(eps_val)
                        if not math.isfinite(eps_f) or eps_f <= 0:
                            continue
                        dt = pd.Timestamp(col)
                        if pd.isna(dt):
                            continue
                        if dt.tzinfo is not None:
                            dt = dt.tz_localize(None)
                        window = hist.loc[: dt]
                        if window.empty:
                            continue
                        px = float(window["Close"].iloc[-1])
                        if not math.isfinite(px) or px <= 0:
                            continue
                        annual_pes.append(px / eps_f)
                    except Exception:
                        continue

        if len(annual_pes) >= 3:
            take = annual_pes[-10:]
            out["估值_十年平均PE"] = round(float(np.mean(take)), 3)
        elif len(annual_pes) > 0:
            notes.append(f"年报有效PE样本仅{len(annual_pes)}个，改试季报")

        if out["估值_十年平均PE"] is None:
            q_avg = _rolling_avg_pe_from_quarterly_eps(tk, hist)
            if q_avg is not None:
                out["估值_十年平均PE"] = round(q_avg, 3)
                notes.append("十年均PE由季报TTM(EPS)滚动近似")
            else:
                if inc is None or inc.empty:
                    notes.append("无年报；季报亦无法推算十年PE")
                elif not any("样本" in n for n in notes):
                    notes.append("年报样本不足且季报无法推算十年PE")

        if notes:
            out["估值_数据说明"] = "；".join(notes)
        _valuation_fill_percentage_vs_avg(out)
    except Exception as ex:
        out["估值_数据说明"] = ("；".join(notes) + "；" if notes else "") + str(ex)[:140]
        _valuation_fill_percentage_vs_avg(out)
    return out


def run_dca_to_now(df: pd.DataFrame, monthly_invest: float = 1000) -> pd.DataFrame:
    """
    从每一个月份作为定投开始日，定投到最新月份。
    """
    df = df.sort_values("date").reset_index(drop=True)
    results = []

    for start_idx in range(len(df) - 12):  # 至少跑满 1 年
        sub = df.iloc[start_idx:].copy()
        dates = sub["date"].tolist()
        prices = sub["price"].astype(float).tolist()

        units = 0.0
        total_invest = 0.0
        values = []
        cashflows = []

        for d, p in zip(dates, prices):
            buy_units = monthly_invest / p
            units += buy_units
            total_invest += monthly_invest
            values.append(units * p)
            cashflows.append((pd.Timestamp(d), -monthly_invest))

        final_date = pd.Timestamp(dates[-1])
        final_value = values[-1]
        cashflows.append((final_date, final_value))

        peak = pd.Series(values).cummax()
        drawdown = pd.Series(values) / peak - 1
        max_loss = drawdown.min()

        total_ret = final_value / total_invest - 1
        rate = xirr(cashflows)

        results.append({
            "start_date": pd.Timestamp(dates[0]).date(),
            "end_date": final_date.date(),
            "years": round((final_date - pd.Timestamp(dates[0])).days / 365.25, 1),
            "months": len(sub),
            "total_invest": round(total_invest, 2),
            "final_value": round(final_value, 2),
            "total_return": total_ret,
            "xirr": rate,
            "max_loss": max_loss,
        })

    return pd.DataFrame(results)


def _horizon_return_stats(
    sub: pd.DataFrame,
    target_years: float,
    band: float = 0.51,
    label: str = "",
) -> Dict[str, Optional[float]]:
    """
    在「起始定投日至回测结束日」约等于 target_years 年的样本上，统计年化 XIRR：
    平均、最高、最低、中位数（及样本数、持有年数中位数）。
    label 建议为「持有期约5年」「持有期约10年」等，直接出现在 Excel 表头中。
    band：年数容差（±0.51 约覆盖 12 个月粒度下的一年 band）。
    窗口内无样本时，用 years 最接近 target_years 的一行作为单点估计。
    """
    if sub.empty or "years" not in sub.columns:
        return {}

    lo, hi = target_years - band, target_years + band
    win = sub[(sub["years"] >= lo) & (sub["years"] <= hi)]

    if win.empty:
        idx = (sub["years"] - target_years).abs().idxmin()
        one = sub.loc[[idx]]
        xr = one["xirr"].iloc[0]
        yr = one["years"].iloc[0]
        return {
            f"{label}_样本数": 1,
            f"{label}_年化XIRR_平均": xr,
            f"{label}_年化XIRR_最高": xr,
            f"{label}_年化XIRR_最低": xr,
            f"{label}_年化XIRR_中位数": xr,
            f"{label}_实际持有年数_中位数": yr,
        }

    return {
        f"{label}_样本数": len(win),
        f"{label}_年化XIRR_平均": win["xirr"].mean(),
        f"{label}_年化XIRR_最高": win["xirr"].max(),
        f"{label}_年化XIRR_最低": win["xirr"].min(),
        f"{label}_年化XIRR_中位数": win["xirr"].median(),
        f"{label}_实际持有年数_中位数": win["years"].median(),
    }


def calc_stats(runs: pd.DataFrame) -> Dict[str, Optional[float]]:
    if runs.empty or "xirr" not in runs:
        return {}

    sub = runs.dropna(subset=["xirr"])
    if sub.empty:
        return {}

    out: Dict[str, Optional[float]] = {
        # 以下为「任意历史起始月」定投至数据最新月末的全体路径统计（持有长短不一）
        "全区间回测_路径数": len(sub),
        "全区间回测_年化XIRR_中位数": sub["xirr"].median(),
        "全区间回测_年化XIRR_平均值": sub["xirr"].mean(),
        "全区间回测_年化XIRR_最低": sub["xirr"].min(),
        "全区间回测_年化XIRR_最高": sub["xirr"].max(),
        "全区间回测_年化XIRR_标准差": sub["xirr"].std(),
        "全区间回测_XIRR为负的路径占比": (sub["xirr"] < 0).mean(),
        "全区间回测_最大回撤_中位数": sub["max_loss"].median(),
        "全区间回测_最大回撤_最差": sub["max_loss"].min(),
    }

    # 持有期约 5 年 / 约 10 年（仅统计「首次买入至回测截止日」落在该年限带内的路径）
    out.update(_horizon_return_stats(sub, 5.0, label="持有期约5年"))
    out.update(_horizon_return_stats(sub, 10.0, label="持有期约10年"))

    return out


def _fmt_pct_for_console(v: Any) -> str:
    """将小数收益率格式化为终端用的百分比字符串。"""
    if v is None:
        return "N/A"
    try:
        x = float(v)
        if not math.isfinite(x):
            return "N/A"
        return f"{x * 100:.2f}%"
    except (TypeError, ValueError):
        return "N/A"
    

def print_horizon_summary(name: str, ticker: str, stats: Dict[str, Any]) -> None:
    """边跑边打印：约 5 年 / 约 10 年持有期带内，年化 XIRR（平均）。"""
    y5 = stats.get("持有期约5年_年化XIRR_平均")
    y10 = stats.get("持有期约10年_年化XIRR_平均")
    print(
        f"  {name} : 持有期约5年(平均XIRR) {_fmt_pct_for_console(y5)}  |  "
        f"持有期约10年(平均XIRR) {_fmt_pct_for_console(y10)}  ({ticker})"
    )


def _str_display_width_units(s: str) -> float:
    """近似「等宽字符」宽度：中文略宽于 ASCII，用于估算 openpyxl 列宽。"""
    w = 0.0
    for c in s:
        w += 2.15 if ord(c) > 127 else 1.05
    return w


def _autofit_cell_width(value: Any, header: str = "") -> float:
    """根据单元格值与列标题估算展示宽度（openpyxl column width 量级）。"""
    h = header or ""
    if value is None:
        return 0.0
    if isinstance(value, (datetime, date)):
        return 12.0
    if isinstance(value, bool):
        return _str_display_width_units(str(value))
    if isinstance(value, (int, float)):
        if any(k in h for k in ("定投起始日", "回测截止日", "start_date", "end_date")):
            return 12.0
        if isinstance(value, float) and any(
            k in h
            for k in [
                "收益率",
                "XIRR",
                "亏损",
                "回撤",
                "回报",
                "xirr",
                "return",
                "loss",
                "概率",
                "占比",
                "参考",
                "差额",
            ]
        ):
            return 12.0
        if isinstance(value, (int, float)):
            s = f"{value:.6g}" if isinstance(value, float) else str(value)
            return min(_str_display_width_units(s) + 1.0, 18.0)
    return min(_str_display_width_units(str(value)) + 1.2, 60.0)


def _find_header_column(ws: Any, must_contain: str) -> int:
    """返回首行表头包含指定子串的列号（1-based），找不到为 0。"""
    for col in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=col).value or "")
        if must_contain in h:
            return col
    return 0


def _sheet_is_dca_detail(ws: Any) -> bool:
    """是否为「逐月起投」明细表（含定投起始日与年化 XIRR 列）。"""
    return bool(
        _find_header_column(ws, "定投起始日") and _find_header_column(ws, "年化XIRR")
    )


def _add_dca_xirr_bar_chart(ws: Any) -> None:
    """
    在明细表右侧插入柱状图：横轴为定投起始日，纵轴为对应行的年化 XIRR。
    仅用于「回测明细」类 sheet（首行含中文列名）。
    """
    if ws.max_row < 2:
        return
    cat_col = _find_header_column(ws, "定投起始日")
    val_col = _find_header_column(ws, "年化XIRR")
    if cat_col < 1 or val_col < 1:
        return

    last_row = ws.max_row
    start_row = 2
    if last_row - start_row + 1 > DCA_BAR_CHART_MAX_POINTS:
        start_row = last_row - DCA_BAR_CHART_MAX_POINTS + 1

    n_pts = last_row - start_row + 1
    chart = BarChart()
    chart.type = "col"
    chart.title = f'{ws.title} — 定投不同起始月 × 年化XIRR（共 {n_pts} 根柱）'
    chart.y_axis.title = "年化XIRR"
    chart.x_axis.title = "定投起始日"
    chart.height = 14  # cm 量级，openpyxl 默认单位
    chart.width = max(20.0, min(34.0, 14.0 + n_pts * 0.07))
    chart.legend = None

    values = Reference(ws, min_col=val_col, min_row=start_row, max_row=last_row)
    chart.add_data(values, titles_from_data=False)
    cats = Reference(ws, min_col=cat_col, min_row=start_row, max_row=last_row)
    chart.set_categories(cats)

    anchor_col = ws.max_column + 2
    ws.add_chart(chart, f"{get_column_letter(anchor_col)}2")


def _add_summary_xirr_comparison_chart(ws: Any) -> None:
    """汇总表：各资产「持有期约5年 / 约10年」年化 XIRR（平均）簇状柱形图。"""
    if ws.max_row < 2:
        return
    name_col = _find_header_column(ws, "资产")
    y5_col = _find_header_column(ws, "持有期约5年_年化XIRR_平均")
    y10_col = _find_header_column(ws, "持有期约10年_年化XIRR_平均")
    if name_col < 1 or y5_col < 1 or y10_col < 1:
        return

    last_row = ws.max_row
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "各资产：持有期约5年 / 约10年 年化XIRR（平均）"
    chart.y_axis.title = "年化XIRR"
    chart.x_axis.title = "资产"
    chart.height = 11
    chart.width = max(16.0, min(30.0, 12.0 + (last_row - 1) * 1.15))

    chart.add_data(
        Reference(ws, min_col=y5_col, min_row=1, max_row=last_row),
        titles_from_data=True,
    )
    chart.add_data(
        Reference(ws, min_col=y10_col, min_row=1, max_row=last_row),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(ws, min_col=name_col, min_row=2, max_row=last_row)
    )

    anchor_col = ws.max_column + 2
    ws.add_chart(chart, f"{get_column_letter(anchor_col)}2")


def _add_benchmark_reference_chart(ws: Any) -> None:
    """指数基准参考表：外部五年/十年年化簇状图；类别多则用条形图更易读。"""
    if ws.max_row < 2:
        return
    name_col = _find_header_column(ws, "指数名称")
    y5_col = _find_header_column(ws, "外部参考表_五年年化")
    y10_col = _find_header_column(ws, "外部参考表_十年年化")
    if name_col < 1 or y5_col < 1 or y10_col < 1:
        return

    last_row = ws.max_row
    n_cat = last_row - 1
    chart = BarChart()
    chart.type = "bar" if n_cat > 12 else "col"
    chart.grouping = "clustered"
    chart.title = "外部参考表：五年 / 十年年化"
    chart.y_axis.title = "年化"
    chart.x_axis.title = "指数"

    chart.height = max(11.0, min(34.0, 10.0 + n_cat * 0.55))
    chart.width = 24.0

    chart.add_data(
        Reference(ws, min_col=y5_col, min_row=1, max_row=last_row),
        titles_from_data=True,
    )
    chart.add_data(
        Reference(ws, min_col=y10_col, min_row=1, max_row=last_row),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(ws, min_col=name_col, min_row=2, max_row=last_row)
    )

    anchor_col = ws.max_column + 2
    ws.add_chart(chart, f"{get_column_letter(anchor_col)}2")


def _autofit_worksheet_columns(
    ws: Any,
    min_width: float = 9.0,
    max_width: float = 72.0,
) -> None:
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col).value or "")
        best = max(
            _autofit_cell_width(header, ""),
            _str_display_width_units(header) + 1.5,
        )
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            best = max(best, _autofit_cell_width(val, header))
        ws.column_dimensions[get_column_letter(col)].width = max(
            min_width, min(best + 2.0, max_width)
        )


def style_excel(path: str) -> None:
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    sub_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="AAAAAA")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                header = str(ws.cell(row=1, column=col).value or "")
                val = cell.value
                if val is None:
                    continue

                # 明细表中的「日期」列：Excel 常以序列号(整数/浮点)存储，需设日期格式才显示为日历
                if any(
                    k in header
                    for k in ("定投起始日", "回测截止日", "start_date", "end_date")
                ):
                    cell.number_format = "yyyy-mm-dd"
                    continue

                if isinstance(val, (datetime, date)):
                    cell.number_format = "yyyy-mm-dd"
                    continue

                if isinstance(val, float) and header.startswith("估值_"):
                    if "百分比" in header:
                        cell.number_format = "0.00"
                    else:
                        cell.number_format = "0.00"
                    continue

                if isinstance(val, float):
                    if any(
                        k in header
                        for k in [
                            "收益率",
                            "XIRR",
                            "亏损",
                            "回撤",
                            "回报",
                            "xirr",
                            "return",
                            "loss",
                            "概率",
                            "占比",
                            "参考",
                            "差额",
                        ]
                    ):
                        cell.number_format = "0.0%"
                    else:
                        cell.number_format = "0.00"

        _autofit_worksheet_columns(ws)

    for ws in wb.worksheets:
        try:
            if ws.title == "汇总":
                _add_summary_xirr_comparison_chart(ws)
            elif ws.title == "指数基准参考":
                _add_benchmark_reference_chart(ws)
            elif _sheet_is_dca_detail(ws):
                _add_dca_xirr_bar_chart(ws)
        except Exception as e:
            print(f"  [图表] 工作表「{ws.title}」：{e}", flush=True)

    if "汇总" in wb.sheetnames:
        ws = wb["汇总"]
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=1).fill = sub_fill
            ws.cell(row=row, column=2).fill = sub_fill

    wb.save(path)


def _process_one_asset(
    name: str,
    info: Dict[str, str],
    bench_lookup: Dict[str, Tuple[Optional[float], Optional[float]]],
) -> Tuple[Dict[str, Any], pd.DataFrame]:
    """下载、回测并生成汇总行与明细 DataFrame。失败则抛异常。"""
    ticker = info["ticker"]
    category = info["category"]
    price_df = download_monthly_price(ticker, START_DATE)
    runs = run_dca_to_now(price_df, MONTHLY_INVEST)
    stats = calc_stats(runs)
    print_horizon_summary(name, ticker, stats)
    row: Dict[str, Any] = {
        "分类": category,
        "资产": name,
        "Ticker": ticker,
        **valuation_metrics_for_ticker(ticker),
        **stats,
    }
    attach_benchmark_columns(row, info, bench_lookup)
    return row, runs


def main() -> None:
    bench_lookup = benchmark_yield_lookup()
    pending: List[Tuple[str, Dict[str, str]]] = list(ASSETS.items())
    # 资产名 -> (汇总行, 明细 runs 或失败时为 None)
    results: Dict[str, Tuple[Dict[str, Any], Optional[pd.DataFrame]]] = {}

    for round_idx in range(DOWNLOAD_MAX_RETRIES + 1):
        if not pending:
            break
        if round_idx > 0:
            delay = min(30.0, 2.0 ** round_idx)
            print(
                f"\n--- 重试第 {round_idx}/{DOWNLOAD_MAX_RETRIES} 轮，"
                f"待处理 {len(pending)} 个标的，等待 {delay:.0f}s ---",
                flush=True,
            )
            time.sleep(delay)

        still_pending: List[Tuple[str, Dict[str, str]]] = []
        for name, info in pending:
            ticker = info["ticker"]
            tag = " （重试）" if round_idx else ""
            print(f"正在处理：{name} / {ticker}{tag}", flush=True)
            try:
                summary_row, runs = _process_one_asset(name, info, bench_lookup)
                results[name] = (summary_row, runs)
            except Exception as e:
                print(f"  {name} / {ticker} ：失败 — {e}", flush=True)
                if round_idx >= DOWNLOAD_MAX_RETRIES:
                    err_row: Dict[str, Any] = {
                        "分类": info["category"],
                        "资产": name,
                        "Ticker": ticker,
                        **valuation_metrics_for_ticker(ticker),
                        "错误": str(e),
                    }
                    attach_benchmark_columns(err_row, info, bench_lookup)
                    results[name] = (err_row, None)
                else:
                    still_pending.append((name, info))
        pending = still_pending

    summary_rows: List[Dict[str, Any]] = []
    detail_sheets: Dict[str, pd.DataFrame] = {}
    for name in ASSETS:
        if name not in results:
            continue
        row, runs = results[name]
        summary_rows.append(row)
        if runs is not None:
            detail_sheets[name[:28]] = runs

    summary = pd.DataFrame(summary_rows)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="汇总", index=False)
        benchmark_reference_frame().to_excel(writer, sheet_name="指数基准参考", index=False)

        for sheet_name, runs in detail_sheets.items():
            safe_name = sheet_name.replace("/", "_").replace("\\", "_")
            runs.rename(columns=RUNS_DETAIL_COLUMNS_ZH).to_excel(writer, sheet_name=safe_name, index=False)

    style_excel(OUTPUT_FILE)

    print(f"\n完成：{OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()
